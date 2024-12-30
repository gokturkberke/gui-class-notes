# Install ttkbootstrap: https://ttkbootstrap.readthedocs.io/en/latest/
# Install openpyxl: https://openpyxl.readthedocs.io/en/stable/
# Install matplotlib: https://matplotlib.org/stable/
# Run database_setup.py first.
# To-Do app with import/export and data visualization functionalities.

import ttkbootstrap as ttk
import dblib
from tkinter import filedialog as fd
from ttkbootstrap.dialogs import Messagebox as msg
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt


class ToDoList(ttk.Window):
    def __init__(self):
        super().__init__(themename="superhero")
        self.geometry("600x720")
        self.title("To-Do List App")
        self.db = dblib.ToDoDatabase()
        self.status_var = ttk.IntVar()
        self.create_widgets()
        self.create_layout()
        self.mainloop()

    def create_widgets(self):
        # Widgets for adding new To-Do items
        self.frm_add = ttk.Labelframe(self, text="Add To-Do Item")
        self.entry_item = ttk.Entry(self.frm_add)
        self.btn_add = ttk.Button(self.frm_add, text="Add", command=self.add_item)

        # Widgets for listing To-Do items
        self.frm_list = ttk.Labelframe(self, text="To-Do List")
        self.tv = ttk.Treeview(self.frm_list, columns=("Item", "Status"), show="headings", selectmode="browse")
        self.tv.heading("Item", text="Item")
        self.tv.heading("Status", text="Status")

        # Widgets for deleting/updating To-Do items
        self.frm_actions = ttk.Labelframe(self, text="Actions")
        self.btn_delete = ttk.Button(self.frm_actions, text="Delete", command=self.delete_item)
        self.cbtn_status = ttk.Checkbutton(self.frm_actions, text="Mark as Completed", variable=self.status_var, command=self.toggle_status, bootstyle="round-toggle")

        # Event bindings for the Treeview and Entry widgets
        self.tv.bind("<<TreeviewSelect>>", self.on_item_select)
        self.entry_item.bind("<Return>", lambda e : self.add_item())

        # Event bindings for the import/export functionalities
        self.bind("<Control-i>", lambda e : self.import_data())
        self.bind("<Control-e>", lambda e : self.export_data())

        # Event binding for pie chart view
        self.bind("<Control-Shift-C>", lambda e : self.show_pie_chart())

        # Populate the Treeview
        self.refresh_list()

    def create_layout(self):
        # Place Labelframe widgets
        self.frm_add.pack(fill="x", padx=15, pady=(15, 0))
        self.frm_list.pack(fill="both", expand=True, padx=15, pady=(15, 0))
        self.frm_actions.pack(fill="x", padx=15, pady=15)

        # Place entry and add button widgets
        self.entry_item.pack(side="left", padx=(15, 0), pady=15, fill="x", expand=True)
        self.btn_add.pack(side="right", padx=15, pady=15)

        # Place the Treeview widget
        self.tv.pack(fill="both", expand=True, padx=15, pady=15)

        # Place delete button and status checkbutton widgets
        self.btn_delete.pack(side="left", padx=15, pady=15)
        self.cbtn_status.pack(side="right", padx=15, pady=15)

    def enable_actions(self):
        self.btn_delete.configure(state="normal")
        self.cbtn_status.configure(state="normal")

    def disable_actions(self):
        self.btn_delete.configure(state="disabled")
        self.cbtn_status.configure(state="disabled")

    def on_item_select(self, event):
        selected_item = self.tv.selection()
        if selected_item:
            # Get the selected item's status
            item_id = int(selected_item[0])
            # print(self.tv.item(item_id)["values"])
            item_status = self.tv.item(item_id)["values"][1]
            # Update the Checkbutton based on the status
            if item_status == "Done":
                self.status_var.set(1)
            else:
                self.status_var.set(0)

    def refresh_list(self):
        # Clear the Treeview content
        for item in self.tv.get_children():
            self.tv.delete(item)

        # Fetch updated data and populate the Treeview
        for row in self.db.get_items():
            self.tv.insert(parent="", index="end", iid=row[0], values=(row[1], row[2]))

        # Enable or disable action buttons based on whether the Treeview has items
        tv_items = self.tv.get_children()
        if tv_items:
            self.enable_actions()
            self.tv.selection_set(tv_items[-1])  # Select the last item
        else:
            self.disable_actions()

    def add_item(self):
        item = self.entry_item.get().strip()
        if item:
            self.db.add_item(item)
            self.refresh_list()
            self.entry_item.delete(0, "end")
            self.entry_item.focus_set()

    def delete_item(self):
        selected_item = self.tv.selection()
        if selected_item:
            item_id = int(selected_item[0])
            self.db.delete_item(item_id)
            self.refresh_list()

    def toggle_status(self):
        selected_item = self.tv.selection()
        if selected_item:
            item_id = int(selected_item[0])
            if self.status_var.get() == 1:
                new_status = "Done"
            else:
                new_status = "Pending"
            self.db.update_status(item_id, new_status)
            self.refresh_list()
            self.tv.selection_set(selected_item) # Keep the current item selected

    def import_data(self):
        file_filter = (("Excel files", "*.xlsx"),
                       ("All files", "*.*"))
        source_file = fd.askopenfile(title="Choose file", filetypes=file_filter)

        if source_file is not None:
            # Load the source file
            wb = load_workbook(source_file.name)
            ws1 = wb.active
            skip_first = True

            # Read the file content
            for row in ws1.iter_rows(values_only=True):
                if skip_first: # Skip the header row
                    skip_first = False
                    continue

                # Get the first two columns
                item, status = row[0], row[1]
                if item:
                    # Insert the item to the database
                    self.db.add_item(item)

                    # Update the item status to "Done" only if the current status is "Done".
                    # No action is taken otherwise as the default status is "Pending" in the database level. 
                    if status == "Done":
                        self.db.update_status(self.db.get_items()[-1][0], "Done")

            # Re-populate the Treeview
            self.refresh_list()

    def export_data(self):
        if not self.tv.get_children():
            msg.show_error(title="Export Data", message="No data available to export.")
            return

        wb = Workbook()
        ws = wb.active

        # Add column headers
        ws.cell(row=1, column=1, value="Item")
        ws.cell(row=1, column=2, value="Status")

        # Read data from Treeview
        row_count = 2
        for item_id in self.tv.get_children():
            # print(self.tv.item(item_id)["values"])
            ws.cell(row=row_count, column=1, value=self.tv.item(item_id)["values"][0])
            ws.cell(row=row_count, column=2, value=self.tv.item(item_id)["values"][1])
            row_count += 1

        # Add summary calculations
        total_items_row = row_count + 1
        completed_items_row = row_count + 2
        not_completed_items_row = row_count + 3

        ws.cell(row=total_items_row, column=1, value="Total items:")
        ws.cell(row=total_items_row, column=2, value=f"=COUNTA(A2:A{row_count - 1})")

        ws.cell(row=completed_items_row, column=1, value="Completed items:")
        ws.cell(row=completed_items_row, column=2, value=f"=COUNTIF(B2:B{row_count - 1}, \"Done\")")

        ws.cell(row=not_completed_items_row, column=1, value="Not completed items:")
        ws.cell(row=not_completed_items_row, column=2, value=f"=COUNTIF(B2:B{row_count - 1}, \"Pending\")")

        # Save the workbook
        wb.save("exported_list.xlsx")
        msg.show_info(title="Export Data", message="The list has been exported and saved as 'exported.xlsx'.")

    def show_pie_chart(self):
        # Fetch data for the chart
        items = self.db.get_items()

        # Exit if no data available to plot
        if len(items) == 0:
            return

        completed = sum(1 for item in items if item[2] == "Done")
        not_completed = sum(1 for item in items if item[2] == "Pending")

        # Prepare pie chart data and settings
        labels = ["Completed", "Not Completed"]
        sizes = [completed, not_completed]
        colors = ["lightblue", "peachpuff"]
        explode = [0.1, 0]  # Highlight the "Completed" section

        # Create the pie chart
        plt.pie(sizes, explode=explode, labels=labels, colors=colors,
                autopct="%.1f%%", shadow=True, startangle=90)
        plt.title("To-Do List Status")
        plt.show()

app = ToDoList()